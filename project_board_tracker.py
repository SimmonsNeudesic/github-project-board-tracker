#!/usr/bin/env python3
"""
GitHub Project Board Tracker

A Python script that exports a full list of issues from a GitHub project board view
and generates project status reports for stakeholders in markdown, CSV, or Excel format.
"""

import os
import sys
import argparse
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
import requests
from field_extractor import AIFieldExtractor


class GitHubProjectBoardTracker:
    """Fetch and export issues from GitHub project boards."""
    
    def __init__(self, token: str, owner: str, repo: Optional[str] = None, 
                 ai_extractor: Optional[AIFieldExtractor] = None,
                 include_pr: bool = False):
        """
        Initialize the tracker.
        
        Args:
            token: GitHub personal access token
            owner: Repository owner (organization or user)
            repo: Repository name (optional, for org-level projects)
            ai_extractor: Optional AI field extractor for --extract-from-body
            include_pr: Whether to include PRs in report output
        """
        self.token = token
        self.owner = owner
        self.repo = repo
        self.ai_extractor = ai_extractor
        self.include_pr = include_pr
        self.headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json',
            'X-GitHub-Api-Version': '2022-11-28'
        }
        self.graphql_url = 'https://api.github.com/graphql'
        self.rest_url = 'https://api.github.com/graphql'
    
    def fetch_project_board_issues(self, project_number: int, view_number: Optional[int] = None, filter_query: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Fetch issues from a GitHub project board, optionally filtered by view or query.
        
        Args:
            project_number: The project board number
            view_number: Optional view number to filter items (deprecated - use filter_query instead)
            filter_query: Optional GitHub search query to filter items (e.g., 'milestone:"Release 1.10.5.8" -is:pr')
            
        Returns:
            List of issues with all required data
        """
        if filter_query:
            return self.fetch_filtered_project_issues(project_number, filter_query)
        elif view_number:
            return self.fetch_project_view_issues(project_number, view_number)
        else:
            return self.fetch_all_project_issues(project_number)
    
    def fetch_all_project_issues(self, project_number: int) -> List[Dict[str, Any]]:
        """
        Fetch all issues from a GitHub project board using GraphQL API.
        
        Args:
            project_number: The project board number
            
        Returns:
            List of issues with all required data
        """
        query = """
        query($owner: String!, $number: Int!, $cursor: String) {
          organization(login: $owner) {
            projectV2(number: $number) {
              title
              items(first: 100, after: $cursor) {
                pageInfo {
                  hasNextPage
                  endCursor
                }
                nodes {
                  id
                  type
                  fieldValues(first: 20) {
                    nodes {
                      ... on ProjectV2ItemFieldTextValue {
                        text
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldNumberValue {
                        number
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldDateValue {
                        date
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldSingleSelectValue {
                        name
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                    }
                  }
                  content {
                    ... on Issue {
                      number
                      title
                      url
                      state
                      createdAt
                      updatedAt
                      body
                      labels(first: 10) {
                        nodes {
                          name
                        }
                      }
                      assignees(first: 5) {
                        nodes {
                          login
                        }
                      }
                    }
                    ... on PullRequest {
                      number
                      title
                      url
                      state
                      createdAt
                      updatedAt
                      body
                      commits(last: 1) {
                        nodes {
                          commit {
                            oid
                          }
                        }
                      }
                    }
                    ... on DraftIssue {
                      title
                      body
                      createdAt
                      updatedAt
                    }
                  }
                }
              }
            }
          }
        }
        """
        
        all_items = []
        has_next_page = True
        cursor = None
        
        while has_next_page:
            variables = {
                'owner': self.owner,
                'number': project_number,
                'cursor': cursor
            }
            
            response = requests.post(
                self.graphql_url,
                json={'query': query, 'variables': variables},
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            if 'errors' in data:
                raise Exception(f"GraphQL errors: {data['errors']}")
            
            project_data = data['data']['organization']['projectV2']
            items = project_data['items']
            all_items.extend(items['nodes'])
            
            has_next_page = items['pageInfo']['hasNextPage']
            cursor = items['pageInfo']['endCursor']
        
        return self._process_items(all_items)
    
    def fetch_project_view_issues(self, project_number: int, view_number: int) -> List[Dict[str, Any]]:
        """
        Fetch issues from a specific view of a GitHub project board.
        Uses the updated GraphQL API to fetch items filtered by view.
        
        Args:
            project_number: The project board number
            view_number: The view number to filter items
            
        Returns:
            List of issues with all required data from the specific view
        """
        query = """
        query($owner: String!, $number: Int!, $cursor: String) {
          organization(login: $owner) {
            projectV2(number: $number) {
              title
              views(first: 50) {
                nodes {
                  id
                  name
                  number
                }
              }
              items(first: 100, after: $cursor) {
                pageInfo {
                  hasNextPage
                  endCursor
                }
                nodes {
                  id
                  type
                  isArchived
                  fieldValues(first: 20) {
                    nodes {
                      ... on ProjectV2ItemFieldTextValue {
                        text
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldNumberValue {
                        number
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldDateValue {
                        date
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldSingleSelectValue {
                        name
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                    }
                  }
                  content {
                    ... on Issue {
                      title
                      url
                      state
                      number
                      createdAt
                      updatedAt
                      author {
                        login
                      }
                      assignees(first: 5) {
                        nodes {
                          login
                        }
                      }
                      labels(first: 10) {
                        nodes {
                          name
                        }
                      }
                      repository {
                        name
                        owner {
                          login
                        }
                      }
                    }
                    ... on PullRequest {
                      title
                      url
                      state
                      number
                      createdAt
                      updatedAt
                      author {
                        login
                      }
                      repository {
                        name
                        owner {
                          login
                        }
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """
        
        # First, get all items and available views
        all_items = []
        has_next_page = True
        cursor = None
        views_info = None
        
        while has_next_page:
            variables = {
                'owner': self.owner,
                'number': project_number,
                'cursor': cursor
            }
            
            response = requests.post(
                self.graphql_url,
                json={'query': query, 'variables': variables},
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            if 'errors' in data:
                raise Exception(f"GraphQL errors: {data['errors']}")
            
            project_data = data['data']['organization']['projectV2']
            
            # Store views info on first iteration
            if views_info is None:
                views_info = project_data['views']['nodes']
                view_exists = any(view['number'] == view_number for view in views_info)
                if not view_exists:
                    available_views = [f"{view['number']} ({view['name']})" for view in views_info]
                    raise Exception(f"View {view_number} not found. Available views: {', '.join(available_views)}")
            
            items = project_data['items']
            all_items.extend(items['nodes'])
            
            has_next_page = items['pageInfo']['hasNextPage']
            cursor = items['pageInfo']['endCursor']
        
        # Note: Since GitHub's API doesn't provide direct view filtering in GraphQL,
        # we return all items. In a real implementation, you would need additional
        # filtering logic based on the view's configuration.
        print(f"Note: Retrieved all {len(all_items)} items from project {project_number}")
        print(f"View {view_number} filtering is not directly supported by GitHub's API")
        print("You may need to manually filter based on your view criteria.")
        
        # Filter out archived items as views typically don't show them
        active_items = [item for item in all_items if not item.get('isArchived', False)]
        print(f"Filtered to {len(active_items)} non-archived items")
        
        return self._process_items(active_items)
    
    def fetch_filtered_project_issues(self, project_number: int, filter_query: str) -> List[Dict[str, Any]]:
        """
        Fetch issues from a GitHub project board filtered by a GitHub search query.
        
        Uses GitHub Search API to evaluate the full query syntax, then intersects
        results with the project's items. This supports all GitHub search operators
        without reimplementing query parsing.
        
        Args:
            project_number: The project board number
            filter_query: GitHub search query (e.g., 'milestone:"Release 1.10.5.8" -is:pr')
            
        Returns:
            List of filtered issues with all required data
        """
        print(f"Applying filter via GitHub Search API: {filter_query}")
        
        # Use GitHub Search API to find matching issues/PRs
        search_url = 'https://api.github.com/search/issues'
        per_page = 100
        page = 1
        matched_set = set()  # Set of (owner, repo, number) tuples
        
        try:
            while True:
                params = {
                    'q': filter_query,
                    'per_page': per_page,
                    'page': page
                }
                response = requests.get(
                    search_url,
                    headers=self.headers,
                    params=params,
                    timeout=30
                )
                response.raise_for_status()
                data = response.json()
                
                items = data.get('items', [])
                for item in items:
                    number = item.get('number')
                    # repository_url format: https://api.github.com/repos/owner/repo
                    repo_url = item.get('repository_url', '')
                    if repo_url:
                        parts = repo_url.rstrip('/').split('/')
                        if len(parts) >= 2:
                            owner = parts[-2]
                            repo_name = parts[-1]
                            matched_set.add((owner, repo_name, number))
                
                # Check if there are more pages
                if len(items) < per_page:
                    break
                
                # Protect against Search API 1000 result limit
                total_count = data.get('total_count', 0)
                if page * per_page >= min(1000, total_count):
                    if total_count > 1000:
                        print(f"Warning: Search returned {total_count} results but API limit is 1000")
                    break
                
                page += 1
                
        except requests.HTTPError as e:
            raise Exception(f"GitHub Search API error: {e}")
        
        print(f"Search API matched {len(matched_set)} issues/PRs")
        
        if not matched_set:
            print("No items matched the search query")
            return []
        
        # Fetch all project items and intersect with search results
        print(f"Fetching project {project_number} items...")
        all_project_items = self.fetch_all_project_issues_raw(project_number)
        
        # Filter project items to only those that matched the search
        filtered_items = []
        for item in all_project_items:
            content = item.get('content')
            if not content:
                continue
            
            repo_info = content.get('repository')
            number = content.get('number')
            
            if repo_info and number:
                owner = repo_info.get('owner', {}).get('login')
                repo_name = repo_info.get('name')
                if (owner, repo_name, number) in matched_set:
                    filtered_items.append(item)
        
        print(f"Filtered to {len(filtered_items)} items (from {len(all_project_items)} total project items)")
        
        # Process and return the filtered items
        return self._process_items(filtered_items)
    
    def fetch_all_project_issues_raw(self, project_number: int) -> List[Dict[str, Any]]:
        """
        Fetch all raw items from a project (before processing).
        
        Args:
            project_number: The project board number
            
        Returns:
            List of raw project items
        """
        query = """
        query($owner: String!, $number: Int!, $cursor: String) {
          organization(login: $owner) {
            projectV2(number: $number) {
              title
              items(first: 100, after: $cursor) {
                pageInfo {
                  hasNextPage
                  endCursor
                }
                nodes {
                  id
                  type
                  isArchived
                  fieldValues(first: 20) {
                    nodes {
                      ... on ProjectV2ItemFieldTextValue {
                        text
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldNumberValue {
                        number
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldDateValue {
                        date
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                      ... on ProjectV2ItemFieldSingleSelectValue {
                        name
                        field {
                          ... on ProjectV2FieldCommon {
                            name
                          }
                        }
                      }
                    }
                  }
                  content {
                    ... on Issue {
                      title
                      url
                      state
                      number
                      createdAt
                      updatedAt
                      author {
                        login
                      }
                      assignees(first: 5) {
                        nodes {
                          login
                        }
                      }
                      labels(first: 10) {
                        nodes {
                          name
                        }
                      }
                      milestone {
                        title
                      }
                      repository {
                        name
                        owner {
                          login
                        }
                      }
                    }
                    ... on PullRequest {
                      title
                      url
                      state
                      number
                      createdAt
                      updatedAt
                      author {
                        login
                      }
                      milestone {
                        title
                      }
                      repository {
                        name
                        owner {
                          login
                        }
                      }
                      closingIssuesReferences(first: 10) {
                        nodes {
                          number
                        }
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """
        
        all_items = []
        has_next_page = True
        cursor = None
        
        while has_next_page:
            variables = {
                'owner': self.owner,
                'number': project_number,
                'cursor': cursor
            }
            
            response = requests.post(
                self.graphql_url,
                json={'query': query, 'variables': variables},
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            data = response.json()
            
            if 'errors' in data:
                raise Exception(f"GraphQL errors: {data['errors']}")
            
            project_data = data['data']['organization']['projectV2']
            items = project_data['items']
            all_items.extend(items['nodes'])
            
            has_next_page = items['pageInfo']['hasNextPage']
            cursor = items['pageInfo']['endCursor']
        
        return all_items
    
    def _normalize_field_name(self, field_name: str) -> str:
        """
        Normalize field names to handle variations (spaces, underscores, case).
        
        Args:
            field_name: Original field name from GitHub
            
        Returns:
            Normalized field name (lowercase, underscores instead of spaces)
        """
        return field_name.lower().replace(' ', '_').replace('-', '_')
    
    def _get_field_value(self, field_values: Dict[str, Any], *field_names: str) -> str:
        """
        Get field value by trying multiple field name variations.
        Returns 'N/A' if not found.
        
        Args:
            field_values: Dictionary of field values
            field_names: List of possible field names to try
            
        Returns:
            Field value or 'N/A'
        """
        for name in field_names:
            # Try exact match first
            if name in field_values and field_values[name]:
                return str(field_values[name])
            # Try normalized match
            normalized_name = self._normalize_field_name(name)
            for key, value in field_values.items():
                if self._normalize_field_name(key) == normalized_name and value:
                    return str(value)
        return 'N/A'
    
    def _fetch_pr_reviews(self, pr_number: int, repo_owner: str, repo_name: str) -> Dict[str, Any]:
        """
        Fetch PR review approvals using GitHub REST API.
        
        Args:
            pr_number: Pull request number
            repo_owner: Repository owner
            repo_name: Repository name
            
        Returns:
            Dictionary with approvers and review states
        """
        reviews_url = f'https://api.github.com/repos/{repo_owner}/{repo_name}/pulls/{pr_number}/reviews'
        
        try:
            response = requests.get(
                reviews_url,
                headers=self.headers,
                timeout=30
            )
            response.raise_for_status()
            reviews = response.json()
            
            # Extract unique approvers (users who approved)
            approvers = []
            seen_users = set()
            
            # Process reviews in reverse order (most recent first)
            for review in reversed(reviews):
                user = review.get('user', {}).get('login')
                state = review.get('state')
                
                # Track most recent approval state per user
                if user and user not in seen_users:
                    seen_users.add(user)
                    if state == 'APPROVED':
                        approvers.append(user)
            
            return {
                'approvers': approvers,
                'approval_count': len(approvers),
                'review_count': len(reviews)
            }
            
        except Exception as e:
            print(f"  Warning: Failed to fetch PR reviews for #{pr_number}: {e}")
            return {
                'approvers': [],
                'approval_count': 0,
                'review_count': 0
            }
    
    def _process_items(self, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Process raw project items into structured issue data.
        Fetches PR approvals and matches them to related issues.
        
        Args:
            items: Raw items from GraphQL API
            
        Returns:
            Processed list of issues with standardized fields
        """
        processed_issues = []
        pr_approvals_map = {}  # Maps issue number to list of approvers from PRs
        
        # First pass: Process PRs to extract approval data
        print("Processing pull requests for approval data...")
        for item in items:
            content = item.get('content')
            if not content:
                continue
            
            # Check if this is a PR
            is_pr = 'closingIssuesReferences' in content
            
            if is_pr:
                pr_number = content.get('number')
                repo_info = content.get('repository', {})
                repo_owner = repo_info.get('owner', {}).get('login')
                repo_name = repo_info.get('name')
                
                if pr_number and repo_owner and repo_name:
                    # Fetch PR reviews/approvals
                    pr_reviews = self._fetch_pr_reviews(pr_number, repo_owner, repo_name)
                    approvers = pr_reviews.get('approvers', [])
                    
                    # Map approvals to linked issues
                    closing_issues = content.get('closingIssuesReferences', {}).get('nodes', [])
                    for linked_issue in closing_issues:
                        issue_num = linked_issue.get('number')
                        if issue_num:
                            if issue_num not in pr_approvals_map:
                                pr_approvals_map[issue_num] = set()
                            pr_approvals_map[issue_num].update(approvers)
        
        print(f"Found approval data for {len(pr_approvals_map)} issues from PRs")
        
        # Second pass: Process all items and build issue records
        for item in items:
            content = item.get('content')
            if not content:
                # Draft issue
                content = {'title': 'Draft Issue', 'url': '', 'state': 'Draft'}
            
            # Extract custom field values with normalization
            field_values = {}
            field_values_normalized = {}
            for field_value in item.get('fieldValues', {}).get('nodes', []):
                if not field_value:
                    continue
                field_name = field_value.get('field', {}).get('name', '')
                if not field_name:
                    continue
                
                value = None
                if 'text' in field_value:
                    value = field_value['text']
                elif 'number' in field_value:
                    value = field_value['number']
                elif 'date' in field_value:
                    value = field_value['date']
                elif 'name' in field_value:
                    value = field_value['name']
                
                if value is not None and value != '':
                    field_values[field_name] = value
                    field_values_normalized[self._normalize_field_name(field_name)] = value
                if value is not None and value != '':
                    field_values[field_name] = value
                    field_values_normalized[self._normalize_field_name(field_name)] = value
            
            # Check if this is a PR (has closingIssuesReferences)
            is_pr = 'closingIssuesReferences' in content
            
            # Skip PRs unless include_pr flag is set
            if is_pr and not self.include_pr:
                continue
            
            # Determine source type (Issue, Discussion, or Pull Request)
            source = 'Issue'
            if item.get('type') == 'DRAFT_ISSUE':
                source = 'Discussion'
            elif is_pr:
                source = 'Pull Request'
            
            # Extract commit SHA for PRs
            commit_shas = []
            if 'commits' in content and content['commits']['nodes']:
                commit_shas = [content['commits']['nodes'][0]['commit']['oid'][:7]]  # Short SHA
            commit_sha_str = ', '.join(commit_shas) if commit_shas else 'N/A'
            
            # Extract labels
            labels = []
            if 'labels' in content and content['labels']['nodes']:
                labels = [label['name'] for label in content['labels']['nodes']]
            
            # Extract assignees
            assignees = []
            if 'assignees' in content and content['assignees']['nodes']:
                assignees = [assignee['login'] for assignee in content['assignees']['nodes']]
            
            # Get Issue ID (try ReqID, Req ID, Issue ID, or fall back to issue number)
            issue_id = self._get_field_value(field_values, 'Issue ID', 'ReqID', 'Req ID', 'ID')
            if issue_id == 'N/A' and content.get('number'):
                issue_id = str(content.get('number'))
            
            # Get Owner (try Owner, Risk Owner, Assignees)
            owner = self._get_field_value(field_values, 'Owner', 'Risk Owner')
            if owner == 'N/A' and assignees:
                owner = assignees[0]  # Use first assignee as owner
            
            # Get Risk (separate from Owner)
            risk = self._get_field_value(field_values, 'Risk', 'Risk Level')
            
            # Get Priority (try field first, then labels)
            priority = self._get_field_value(field_values, 'Priority', 'Priority Level')
            if priority == 'N/A':
                priority_label = self._get_priority_from_labels(labels)
                if priority_label:
                    priority = priority_label
            
            # Get Status (try field first, then state)
            status = self._get_field_value(field_values, 'Status', 'State')
            if status == 'N/A':
                status = content.get('state', 'N/A')
            
            # Extract Business Need and Acceptance Criteria from issue body
            business_need = self._get_field_value(field_values, 'Business Need', 'Business_Need')
            if business_need == 'N/A':
                business_need = self._extract_section(content.get('body', ''), 'Business Need') or 'N/A'
            
            acceptance_criteria = self._get_field_value(field_values, 'Acceptance Criteria', 'Acceptance_Criteria')
            if acceptance_criteria == 'N/A':
                acceptance_criteria = self._extract_section(content.get('body', ''), 'Acceptance Criteria') or 'N/A'
            
            # Try AI extraction for missing fields if enabled
            ai_extracted = {}
            if self.ai_extractor and content.get('number'):
                # Identify which fields need extraction (have N/A values)
                needs_extraction = (
                    business_need == 'N/A' or
                    acceptance_criteria == 'N/A' or
                    self._get_field_value(field_values, 'Test Case ID(s)', 'Test Case IDs') == 'N/A' or
                    self._get_field_value(field_values, 'Test Evidence URL', 'Test Evidence') == 'N/A' or
                    self._get_field_value(field_values, 'Design Artifact(s) URL', 'Design Artifacts') == 'N/A' or
                    risk == 'N/A' or
                    self._get_field_value(field_values, 'Release Version', 'Release') == 'N/A' or
                    self._get_field_value(field_values, 'Change Log', 'Changelog') == 'N/A'
                )
                
                if needs_extraction:
                    try:
                        # Extract repository name from issue URL
                        repo_name = self.repo
                        if not repo_name and content.get('url'):
                            # Parse from URL: https://github.com/owner/repo/issues/123
                            url_parts = content['url'].split('/')
                            if len(url_parts) >= 5:
                                repo_name = url_parts[4]
                        
                        if repo_name:
                            ai_extracted = self.ai_extractor.extract_fields(
                                content, self.owner, repo_name
                            )
                    except Exception as e:
                        print(f"  Warning: AI extraction failed for issue #{content.get('number')}: {e}")
            
            # Get approval fields from project fields first
            approval_po = self._get_field_value(field_values, 'Approvals: Product Owner', 'Approval Product', 'Approval_Product', 'Product Owner Approval')
            approval_qa = self._get_field_value(field_values, 'Approvals: QA Lead', 'Approval QA Lead', 'Approval_QA_Lead', 'QA Lead Approval')
            approval_exec = self._get_field_value(field_values, 'Approvals: Exec Sponsor', 'Approval Sponsor', 'Approval_Sponsor', 'Exec Sponsor Approval')
            
            # If issue has linked PRs with approvals, use those approvers
            issue_number = content.get('number')
            if issue_number and issue_number in pr_approvals_map:
                pr_approvers = sorted(pr_approvals_map[issue_number])
                approvers_str = ', '.join(pr_approvers)
                
                # Set approvals if not already filled in
                if approval_po == 'N/A':
                    approval_po = approvers_str
                if approval_qa == 'N/A':
                    approval_qa = approvers_str
                if approval_exec == 'N/A':
                    approval_exec = approvers_str
            
            # Build the issue record with new column names
            # Use AI-extracted values if field is N/A and AI found something
            issue_record = {
                'Issue ID': issue_id,
                'Title': content.get('title', 'N/A'),
                'Source': source,
                'GitHub Issue URL': content.get('url', 'N/A'),
                'Business Need': ai_extracted.get('business_need', business_need),
                'Acceptance Criteria': ai_extracted.get('acceptance_criteria', acceptance_criteria),
                'Design Artifact(s) URL': ai_extracted.get('design_artifacts_url') or self._get_field_value(field_values, 'Design Artifact(s) URL', 'Design Artifacts', 'Design Artifact URLs', 'Design_Artifact_URLs'),
                'Test Case ID(s)': ai_extracted.get('test_case_ids') or self._get_field_value(field_values, 'Test Case ID(s)', 'Test Case IDs', 'Test Case ID', 'Test_Case_ID'),
                'Test Evidence URL': ai_extracted.get('test_evidence_url') or self._get_field_value(field_values, 'Test Evidence URL', 'Test Evidence', 'Test_Evidence_URL'),
                'Linked PR(s) URL': content.get('url', 'N/A') if is_pr else 'N/A',
                'Commit SHA(s)': commit_sha_str,
                'Status': status,
                'Priority': priority,
                'Risk': ai_extracted.get('risk', risk),
                'Owner': owner,
                'Approvals: Product Owner': approval_po,
                'Approvals: QA Lead': approval_qa,
                'Approvals: Exec Sponsor': approval_exec,
                'Release Version': ai_extracted.get('release_version') or self._get_field_value(field_values, 'Release Version', 'Release', 'Version'),
                'Created Date': content.get('createdAt', 'N/A')[:10] if content.get('createdAt') else 'N/A',
                'Updated Date': content.get('updatedAt', 'N/A')[:10] if content.get('updatedAt') else 'N/A',
                'Change Log': ai_extracted.get('change_log') or self._get_field_value(field_values, 'Change Log', 'Changelog', 'Change_Log')
            }
            
            processed_issues.append(issue_record)
        
        return processed_issues
    
    def _extract_section(self, body: str, section_name: str) -> str:
        """Extract a section from issue body markdown."""
        if not body:
            return ''
        
        lines = body.split('\n')
        in_section = False
        section_content = []
        
        for line in lines:
            if section_name.lower() in line.lower() and (line.startswith('#') or line.startswith('**')):
                in_section = True
                continue
            elif in_section and (line.startswith('#') or (line.startswith('**') and '**' in line[2:])):
                break
            elif in_section:
                section_content.append(line)
        
        return '\n'.join(section_content).strip()
    
    def _get_priority_from_labels(self, labels: List[str]) -> str:
        """Extract priority from labels."""
        # Look for common priority patterns
        priority_patterns = ['priority', 'p0', 'p1', 'p2', 'p3', 'high', 'medium', 'low', 'critical']
        for label in labels:
            label_lower = label.lower()
            for pattern in priority_patterns:
                if pattern in label_lower:
                    return label
        return 'N/A'
    
    def export_to_csv(self, issues: List[Dict[str, Any]], output_file: str):
        """
        Export issues to CSV format.
        
        Args:
            issues: List of issue records
            output_file: Path to output CSV file
        """
        if not issues:
            print("No issues to export.")
            return
        
        fieldnames = [
            'Issue ID', 'Title', 'Source', 'GitHub Issue URL', 'Business Need',
            'Acceptance Criteria', 'Design Artifact(s) URL', 'Test Case ID(s)',
            'Test Evidence URL', 'Linked PR(s) URL', 'Commit SHA(s)', 'Status', 'Priority',
            'Risk', 'Owner', 'Approvals: Product Owner', 'Approvals: QA Lead',
            'Approvals: Exec Sponsor', 'Release Version', 'Created Date',
            'Updated Date', 'Change Log'
        ]
        
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(issues)
        
        print(f"Exported {len(issues)} issues to {output_file}")
    
    def export_to_markdown(self, issues: List[Dict[str, Any]], output_file: str):
        """
        Export issues to Markdown format.
        
        Args:
            issues: List of issue records
            output_file: Path to output Markdown file
        """
        if not issues:
            print("No issues to export.")
            return
        
        with open(output_file, 'w', encoding='utf-8') as mdfile:
            mdfile.write("# Project Board Status Report\n\n")
            mdfile.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            mdfile.write(f"Total Issues: {len(issues)}\n\n")
            
            # Summary by status
            status_counts = {}
            for issue in issues:
                status = issue.get('Status', 'Unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            
            mdfile.write("## Status Summary\n\n")
            for status, count in sorted(status_counts.items()):
                mdfile.write(f"- {status}: {count}\n")
            mdfile.write("\n")
            
            # Detailed table
            mdfile.write("## Detailed Issue List\n\n")
            mdfile.write("| Issue ID | Title | Source | Status | Priority | Created | Updated |\n")
            mdfile.write("|----------|-------|--------|--------|----------|---------|----------|\n")
            
            for issue in issues:
                issue_id = self._escape_markdown(str(issue.get('Issue ID', 'N/A')))
                title = self._escape_markdown(issue.get('Title', 'N/A'))
                source = self._escape_markdown(issue.get('Source', 'N/A'))
                status = self._escape_markdown(issue.get('Status', 'N/A'))
                priority = self._escape_markdown(issue.get('Priority', 'N/A'))
                created = issue.get('Created Date', 'N/A')
                updated = issue.get('Updated Date', 'N/A')
                
                mdfile.write(f"| {issue_id} | {title} | {source} | {status} | {priority} | {created} | {updated} |\n")
            
            # Detailed sections
            mdfile.write("\n## Detailed Issue Information\n\n")
            for issue in issues:
                mdfile.write(f"### {issue.get('Issue ID', 'N/A')} - {issue.get('Title', 'Untitled')}\n\n")
                mdfile.write(f"**Source:** {issue.get('Source', 'N/A')}\n\n")
                mdfile.write(f"**URL:** {issue.get('GitHub Issue URL', 'N/A')}\n\n")
                mdfile.write(f"**Status:** {issue.get('Status', 'N/A')}\n\n")
                mdfile.write(f"**Priority:** {issue.get('Priority', 'N/A')}\n\n")
                mdfile.write(f"**Risk:** {issue.get('Risk', 'N/A')}\n\n")
                mdfile.write(f"**Owner:** {issue.get('Owner', 'N/A')}\n\n")
                
                if issue.get('Business Need') and issue.get('Business Need') != 'N/A':
                    mdfile.write(f"**Business Need:** {issue.get('Business Need')}\n\n")
                
                if issue.get('Acceptance Criteria') and issue.get('Acceptance Criteria') != 'N/A':
                    mdfile.write(f"**Acceptance Criteria:** {issue.get('Acceptance Criteria')}\n\n")
                
                if issue.get('Linked PR(s) URL') and issue.get('Linked PR(s) URL') != 'N/A':
                    mdfile.write(f"**Linked PR(s):** {issue.get('Linked PR(s) URL')}\n\n")
                
                if issue.get('Commit SHA(s)') and issue.get('Commit SHA(s)') != 'N/A':
                    mdfile.write(f"**Commit SHA(s):** {issue.get('Commit SHA(s)')}\n\n")
                
                mdfile.write("---\n\n")
        
        print(f"Exported {len(issues)} issues to {output_file}")
    
    def export_to_excel(self, issues: List[Dict[str, Any]], output_file: str):
        """
        Export issues to Excel format.
        
        Args:
            issues: List of issue records
            output_file: Path to output Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("Error: openpyxl library is required for Excel export.")
            print("Install it with: pip install openpyxl")
            return
        
        if not issues:
            print("No issues to export.")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Board Status"
        
        # Define headers
        headers = [
            'Issue ID', 'Title', 'Source', 'GitHub Issue URL', 'Business Need',
            'Acceptance Criteria', 'Design Artifact(s) URL', 'Test Case ID(s)',
            'Test Evidence URL', 'Linked PR(s) URL', 'Commit SHA(s)', 'Status', 'Priority',
            'Risk', 'Owner', 'Approvals: Product Owner', 'Approvals: QA Lead',
            'Approvals: Exec Sponsor', 'Release Version', 'Created Date',
            'Updated Date', 'Change Log'
        ]
        
        # Style for header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, issue in enumerate(issues, 2):
            for col_num, header in enumerate(headers, 1):
                value = issue.get(header, '')
                # Sanitize value for Excel compatibility
                if isinstance(value, str):
                    # Remove or replace problematic characters
                    value = value.replace('\x00', '').replace('\x01', '').replace('\x02', '')
                    # Limit cell content length to prevent Excel issues
                    if len(value) > 32767:  # Excel cell limit
                        value = value[:32764] + "..."
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        wb.save(output_file)
        print(f"Exported {len(issues)} issues to {output_file}")
    
    def _escape_markdown(self, text: str) -> str:
        """Escape special characters in markdown."""
        if not text:
            return ''
        return str(text).replace('|', '\\|').replace('\n', ' ')


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Export GitHub project board issues to various formats',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export to CSV (entire project)
  python project_board_tracker.py --owner myorg --project 1 --format csv --output report.csv

  # Export issues only (no PRs) with filter (recommended)
  python project_board_tracker.py --owner neudesic --project 137 --filter 'milestone:"Release 1.10.5.8" is:issue' --format excel --output report.xlsx

  # Export with PRs included to gather approval data, but exclude PRs from report
  python project_board_tracker.py --owner neudesic --project 137 --filter 'milestone:"Release 1.10.5.8"' --format excel --output report.xlsx

  # Export with PRs included in the report
  python project_board_tracker.py --owner neudesic --project 137 --filter 'milestone:"Release 1.10.5.8"' --include-pr --format excel --output report.xlsx

  # Export with multiple filter criteria
  python project_board_tracker.py --owner myorg --project 1 --filter 'state:open label:bug assignee:john' --format markdown --output bugs.md

  # Export view (deprecated - use --filter instead)
  python project_board_tracker.py --owner myorg --project 1 --view 43 --format excel --output report.xlsx

Environment Variables:
  GITHUB_TOKEN: GitHub personal access token (required)
        """
    )
    
    parser.add_argument(
        '--owner',
        required=True,
        help='Repository owner (organization or user)'
    )
    parser.add_argument(
        '--repo',
        help='Repository name (optional, for org-level projects)'
    )
    parser.add_argument(
        '--project',
        type=int,
        required=True,
        help='Project board number'
    )
    parser.add_argument(
        '--view',
        type=int,
        help='Project view number (optional, deprecated - use --filter instead)'
    )
    parser.add_argument(
        '--filter',
        type=str,
        help='GitHub search query to filter items (e.g., \'milestone:"Release 1.10.5.8" -is:pr\')'
    )
    parser.add_argument(
        '--include-pr',
        action='store_true',
        help='Include pull requests in the report output (default: exclude PRs from reports)'
    )
    parser.add_argument(
        '--format',
        choices=['csv', 'markdown', 'excel'],
        default='csv',
        help='Output format (default: csv)'
    )
    parser.add_argument(
        '--output',
        help='Output file path (default: project_board_report.[format])'
    )
    parser.add_argument(
        '--token',
        help='GitHub personal access token (or set GITHUB_TOKEN environment variable)'
    )
    
    # AI Extraction options
    parser.add_argument(
        '--extract-from-body',
        action='store_true',
        help='Extract missing fields from issue content using AI (Azure OpenAI)'
    )
    parser.add_argument(
        '--azure-openai-endpoint',
        help='Azure OpenAI endpoint URL (or set AZURE_OPENAI_ENDPOINT environment variable)'
    )
    parser.add_argument(
        '--azure-openai-key',
        help='Azure OpenAI API key (or set AZURE_OPENAI_API_KEY environment variable)'
    )
    parser.add_argument(
        '--azure-deployment',
        default='gpt-4.1-mini',
        help='Azure OpenAI deployment name (default: gpt-4.1-mini)'
    )
    parser.add_argument(
        '--cache-dir',
        default='.ai_cache',
        help='Directory for caching AI extractions (default: .ai_cache)'
    )
    
    args = parser.parse_args()
    
    # Get GitHub token
    token = args.token or os.environ.get('GITHUB_TOKEN')
    if not token:
        print("Error: GitHub token is required. Set GITHUB_TOKEN environment variable or use --token")
        sys.exit(1)
    
    # Initialize AI extractor if requested
    ai_extractor = None
    if args.extract_from_body:
        azure_endpoint = args.azure_openai_endpoint or os.environ.get('AZURE_OPENAI_ENDPOINT')
        azure_key = args.azure_openai_key or os.environ.get('AZURE_OPENAI_API_KEY')
        
        if not azure_endpoint or not azure_key:
            print("Error: Azure OpenAI credentials required for --extract-from-body")
            print("  Set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY environment variables")
            print("  Or use --azure-openai-endpoint and --azure-openai-key arguments")
            sys.exit(1)
        
        print(f"Initializing AI field extractor (deployment: {args.azure_deployment})...")
        ai_extractor = AIFieldExtractor(
            endpoint=azure_endpoint,
            api_key=azure_key,
            deployment=args.azure_deployment,
            cache_dir=args.cache_dir,
            github_token=token,
            owner=args.owner,
            project=str(args.project)
        )
    
    # Determine output file
    if not args.output:
        extension = 'csv' if args.format == 'csv' else ('md' if args.format == 'markdown' else 'xlsx')
        args.output = f'project_board_report.{extension}'
    
    try:
        # Initialize tracker
        tracker = GitHubProjectBoardTracker(token, args.owner, args.repo, ai_extractor, args.include_pr)
        
        # Fetch issues
        if args.filter:
            print(f"Fetching issues from project {args.project} with filter: {args.filter}")
            issues = tracker.fetch_project_board_issues(args.project, filter_query=args.filter)
        elif args.view:
            print(f"Fetching issues from project {args.project}, view {args.view}...")
            issues = tracker.fetch_project_board_issues(args.project, view_number=args.view)
        else:
            print(f"Fetching issues from project {args.project} (all items)...")
            issues = tracker.fetch_project_board_issues(args.project)
        print(f"Found {len(issues)} issues")
        
        # Export to requested format
        if args.format == 'csv':
            tracker.export_to_csv(issues, args.output)
        elif args.format == 'markdown':
            tracker.export_to_markdown(issues, args.output)
        elif args.format == 'excel':
            tracker.export_to_excel(issues, args.output)
        
        print("Export completed successfully!")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
