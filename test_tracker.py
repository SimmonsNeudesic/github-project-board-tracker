#!/usr/bin/env python3
"""
Unit tests for project_board_tracker.py

Tests the core functionality of the GitHub Project Board Tracker
without requiring actual GitHub API access.
"""

import os
import sys
import csv
import tempfile
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from project_board_tracker import GitHubProjectBoardTracker


def create_mock_issues():
    """Create mock issue data for testing."""
    return [
        {
            'ReqID': 'REQ-001',
            'Title': 'Test Issue 1',
            'Source': 'Issue',
            'Issue_URL': 'https://github.com/test/repo/issues/1',
            'Business_Need': 'Test business need',
            'Acceptance_Criteria': 'Test criteria',
            'Design_Artifact_URLs': 'https://example.com/design',
            'Test_Case_ID': 'TC-001',
            'Test_Evidence_URL': 'https://example.com/evidence',
            'PR_URL': 'https://github.com/test/repo/pull/1',
            'Commit_SHA': 'abc123',
            'Status': 'In Progress',
            'Priority': 'High',
            'Risk_Owner': 'test.user',
            'Approval_Product': 'Approved',
            'Approval_QA_Lead': 'Pending',
            'Approval_Sponsor': 'Approved',
            'Release_Version': 'v1.0.0',
            'Created_Date': '2024-01-01T00:00:00Z',
            'Update_Date': '2024-01-02T00:00:00Z',
            'Change_Log': 'Initial creation'
        },
        {
            'ReqID': 'REQ-002',
            'Title': 'Test Issue 2',
            'Source': 'Pull Request',
            'Issue_URL': 'https://github.com/test/repo/pull/2',
            'Business_Need': '',
            'Acceptance_Criteria': '',
            'Design_Artifact_URLs': '',
            'Test_Case_ID': '',
            'Test_Evidence_URL': '',
            'PR_URL': 'https://github.com/test/repo/pull/2',
            'Commit_SHA': 'def456',
            'Status': 'Done',
            'Priority': 'Medium',
            'Risk_Owner': '',
            'Approval_Product': '',
            'Approval_QA_Lead': '',
            'Approval_Sponsor': '',
            'Release_Version': '',
            'Created_Date': '2024-01-03T00:00:00Z',
            'Update_Date': '2024-01-04T00:00:00Z',
            'Change_Log': ''
        }
    ]


def test_tracker_initialization():
    """Test that tracker initializes correctly."""
    print("Testing tracker initialization...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner', 'test_repo')
    assert tracker.token == 'test_token'
    assert tracker.owner == 'test_owner'
    assert tracker.repo == 'test_repo'
    print("✓ Tracker initialization test passed")


def test_markdown_escaping():
    """Test markdown special character escaping."""
    print("Testing markdown escaping...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    
    # Test pipe escaping
    result = tracker._escape_markdown('test | with | pipes')
    assert '\\|' in result
    
    # Test newline replacement
    result = tracker._escape_markdown('test\nwith\nnewlines')
    assert '\n' not in result
    
    # Test None handling
    result = tracker._escape_markdown(None)
    assert result == ''
    
    print("✓ Markdown escaping test passed")


def test_priority_extraction():
    """Test priority extraction from labels."""
    print("Testing priority extraction...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    
    labels = ['bug', 'priority:high', 'enhancement']
    priority = tracker._get_priority_from_labels(labels)
    assert priority == 'priority:high'
    
    labels = ['bug', 'enhancement']
    priority = tracker._get_priority_from_labels(labels)
    assert priority == ''
    
    print("✓ Priority extraction test passed")


def test_section_extraction():
    """Test extraction of sections from issue body."""
    print("Testing section extraction...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    
    body = '''
# Business Need
This is the business need section.
It has multiple lines.

## Acceptance Criteria
- Criterion 1
- Criterion 2

## Other Section
Other content
'''
    
    business_need = tracker._extract_section(body, 'Business Need')
    assert 'business need section' in business_need.lower()
    assert 'multiple lines' in business_need.lower()
    
    criteria = tracker._extract_section(body, 'Acceptance Criteria')
    assert 'Criterion 1' in criteria
    assert 'Criterion 2' in criteria
    
    # Test non-existent section
    result = tracker._extract_section(body, 'Non-Existent')
    assert result == ''
    
    # Test None body
    result = tracker._extract_section(None, 'Business Need')
    assert result == ''
    
    print("✓ Section extraction test passed")


def test_csv_export():
    """Test CSV export functionality."""
    print("Testing CSV export...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    issues = create_mock_issues()
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv') as f:
        output_file = f.name
    
    try:
        tracker.export_to_csv(issues, output_file)
        
        # Verify file was created
        assert os.path.exists(output_file)
        
        # Read and verify contents
        with open(output_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            rows = list(reader)
            
            assert len(rows) == 2
            assert rows[0]['ReqID'] == 'REQ-001'
            assert rows[0]['Title'] == 'Test Issue 1'
            assert rows[0]['Status'] == 'In Progress'
            assert rows[1]['ReqID'] == 'REQ-002'
        
        print("✓ CSV export test passed")
    finally:
        if os.path.exists(output_file):
            os.unlink(output_file)


def test_markdown_export():
    """Test Markdown export functionality."""
    print("Testing Markdown export...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    issues = create_mock_issues()
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.md') as f:
        output_file = f.name
    
    try:
        tracker.export_to_markdown(issues, output_file)
        
        # Verify file was created
        assert os.path.exists(output_file)
        
        # Read and verify contents
        with open(output_file, 'r', encoding='utf-8') as mdfile:
            content = mdfile.read()
            
            assert '# Project Board Status Report' in content
            assert 'Total Issues: 2' in content
            assert 'REQ-001' in content
            assert 'REQ-002' in content
            assert 'Test Issue 1' in content
            assert 'Test Issue 2' in content
            assert 'Status Summary' in content
            assert 'In Progress' in content
            assert 'Done' in content
        
        print("✓ Markdown export test passed")
    finally:
        if os.path.exists(output_file):
            os.unlink(output_file)


def test_excel_export():
    """Test Excel export functionality."""
    print("Testing Excel export...")
    try:
        import openpyxl
    except ImportError:
        print("⚠ Skipping Excel test (openpyxl not available)")
        return
    
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    issues = create_mock_issues()
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as f:
        output_file = f.name
    
    try:
        tracker.export_to_excel(issues, output_file)
        
        # Verify file was created
        assert os.path.exists(output_file)
        
        # Read and verify contents
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check headers
        assert ws.cell(1, 1).value == 'ReqID'
        assert ws.cell(1, 2).value == 'Title'
        assert ws.cell(1, 12).value == 'Status'
        
        # Check data
        assert ws.cell(2, 1).value == 'REQ-001'
        assert ws.cell(2, 2).value == 'Test Issue 1'
        assert ws.cell(3, 1).value == 'REQ-002'
        
        print("✓ Excel export test passed")
    finally:
        if os.path.exists(output_file):
            os.unlink(output_file)


def test_empty_issues():
    """Test export functions with empty issue lists."""
    print("Testing empty issues handling...")
    tracker = GitHubProjectBoardTracker('test_token', 'test_owner')
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv') as f:
        output_file = f.name
    
    try:
        # Should not crash with empty list
        tracker.export_to_csv([], output_file)
        tracker.export_to_markdown([], output_file)
        
        print("✓ Empty issues test passed")
    finally:
        if os.path.exists(output_file):
            os.unlink(output_file)


def run_all_tests():
    """Run all tests."""
    print("=" * 60)
    print("Running Project Board Tracker Tests")
    print("=" * 60)
    print()
    
    tests = [
        test_tracker_initialization,
        test_markdown_escaping,
        test_priority_extraction,
        test_section_extraction,
        test_csv_export,
        test_markdown_export,
        test_excel_export,
        test_empty_issues,
    ]
    
    passed = 0
    failed = 0
    
    for test_func in tests:
        try:
            test_func()
            passed += 1
        except Exception as e:
            print(f"✗ {test_func.__name__} failed: {e}")
            failed += 1
        print()
    
    print("=" * 60)
    print(f"Test Results: {passed} passed, {failed} failed")
    print("=" * 60)
    
    return failed == 0


if __name__ == '__main__':
    success = run_all_tests()
    sys.exit(0 if success else 1)
